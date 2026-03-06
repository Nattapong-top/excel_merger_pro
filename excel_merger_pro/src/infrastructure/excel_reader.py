# File: src/infrastructure/excel_reader.py
from typing import List, Any, Iterator
from src.application.interfaces import ISheetReader
from src.domain.value_objects import FilePath, SheetName

class PandasSheetReader(ISheetReader):
    def __init__(self, logger=None):
        """Initialize reader with optional logger"""
        self.logger = logger
    
    def get_sheet_names(self, path: FilePath) -> List[SheetName]:
        import pandas as pd
        try:
            with pd.ExcelFile(path.value) as excel_file:
                return [SheetName(name) for name in excel_file.sheet_names]
        except Exception as e:
            raise ValueError(f"Cannot read file '{path.value}': {e}")

    def read_sheet(self, path: FilePath, sheet_name: SheetName, usecols: List[str] = None) -> Any:
        import pandas as pd
        try:
            # อ่านข้อมูลจาก Sheet ที่ระบุ
            # ลองใช้ engine='calamine' ก่อน (เร็วที่สุด 5-10x)
            # ถ้าไม่มี fallback ไป openpyxl
            # ไม่ใช้ dtype=str ทั้งหมดเพื่อเพิ่มความเร็ว (เร็วขึ้น 50-70%)
            
            # ถ้ามีการระบุ usecols ให้อ่านเฉพาะคอลัมน์นั้น (เร็วขึ้นอีก 50-70%)
            engines_to_try = ['calamine', 'openpyxl']
            last_error = None
            
            for engine in engines_to_try:
                try:
                    # ลองหา header row ที่ถูกต้อง (skip merged cells)
                    header_row = self._find_header_row(path, sheet_name, engine)
                    
                    # ลองอ่านด้วย usecols ก่อน (ถ้ามี)
                    if usecols:
                        try:
                            if self.logger:
                                self.logger.info(f"    Reading only {len(usecols)} selected columns with {engine} engine")
                            df = pd.read_excel(
                                path.value, 
                                sheet_name=sheet_name.value,
                                engine=engine,
                                usecols=usecols,
                                header=header_row
                            )
                        except (ValueError, KeyError) as e:
                            # usecols ไม่ตรงกับ columns ในไฟล์ - อ่านทุก column แทน
                            if self.logger:
                                self.logger.info(f"    Selected columns not found in this file, reading all columns")
                            df = pd.read_excel(
                                path.value, 
                                sheet_name=sheet_name.value,
                                engine=engine,
                                header=header_row
                            )
                    else:
                        if self.logger:
                            self.logger.info(f"    Using {engine} engine for reading")
                        df = pd.read_excel(
                            path.value, 
                            sheet_name=sheet_name.value,
                            engine=engine,
                            header=header_row
                        )
                    
                    # ทำความสะอาด column names (ลบ NaN, Unnamed, ซ้ำกัน)
                    df = self._clean_column_names(df)
                    
                    # แปลงเฉพาะคอลัมน์ที่เป็น object (text) ให้เป็น string
                    # เพื่อป้องกันปัญหา mixed types และ preserve leading zeros
                    for col in df.columns:
                        if df[col].dtype == 'object':
                            df[col] = df[col].astype(str)
                    
                    return df
                    
                except (ImportError, ValueError, Exception) as e:
                    # Engine ไม่มีหรือใช้ไม่ได้ ลองตัวถัดไป
                    last_error = e
                    # ถ้าเป็น calamine error ให้ลอง openpyxl
                    if engine == 'calamine':
                        if self.logger:
                            self.logger.info(f"    Calamine failed, trying openpyxl...")
                        continue
                    else:
                        # openpyxl ก็ล้มเหลว
                        raise
            
            # ถ้าทุก engine ล้มเหลว
            raise ValueError(f"Error reading sheet '{sheet_name.value}': {last_error}")
            
        except Exception as e:
            raise ValueError(f"Error reading sheet '{sheet_name.value}': {e}")
    
    def _find_header_row(self, path: FilePath, sheet_name: SheetName, engine: str) -> int:
        """
        หา header row ที่ถูกต้อง โดย skip แถวที่มี merged cells หรือ title rows
        
        กลยุทธ์:
        1. อ่าน 7 แถวแรก
        2. หาแถวที่มี column names มากที่สุดและไม่ซ้ำกัน
        3. ถ้าไม่เจอ ใช้แถว 0 (default)
        """
        import pandas as pd
        
        try:
            # อ่าน 7 แถวแรกเพื่อหา header
            df_sample = pd.read_excel(
                path.value,
                sheet_name=sheet_name.value,
                engine=engine,
                nrows=7,
                header=None  # อ่านแบบไม่มี header
            )
            
            # หาแถวที่เหมาะสมที่สุด
            best_row = 0
            max_valid_cols = 0
            
            for row_idx in range(min(7, len(df_sample))):
                row_data = df_sample.iloc[row_idx]
                
                # นับจำนวน column ที่ valid (ไม่ใช่ NaN, ไม่ว่าง, ไม่ซ้ำ)
                valid_cols = row_data.dropna()
                valid_cols = valid_cols[valid_cols.astype(str).str.strip() != '']
                
                # ตรวจสอบว่าไม่มี column ซ้ำกัน
                unique_count = len(valid_cols.unique())
                
                if unique_count > max_valid_cols:
                    max_valid_cols = unique_count
                    best_row = row_idx
            
            return best_row
            
        except:
            # ถ้าหาไม่เจอ ใช้แถว 0
            return 0
    
    def _clean_column_names(self, df):
        """
        ทำความสะอาด column names:
        - ลบ NaN
        - ลบ Unnamed columns
        - จัดการ column ที่ซ้ำกัน
        """
        import pandas as pd
        
        new_columns = []
        seen = {}
        
        for col in df.columns:
            # แปลงเป็น string
            col_str = str(col)
            
            # ถ้าเป็น NaN หรือ Unnamed ให้ตั้งชื่อใหม่
            if pd.isna(col) or col_str.startswith('Unnamed:') or col_str == 'nan':
                col_str = f'Column_{len(new_columns)}'
            
            # จัดการ column ที่ซ้ำกัน
            if col_str in seen:
                seen[col_str] += 1
                col_str = f'{col_str}_{seen[col_str]}'
            else:
                seen[col_str] = 0
            
            new_columns.append(col_str)
        
        df.columns = new_columns
        return df
    
    def read_sheet_chunked(
        self, 
        path: FilePath, 
        sheet_name: SheetName, 
        chunk_size: int,
        usecols: List[str] = None
    ) -> Iterator[Any]:
        """
        อ่าน Excel Sheet ทีละ chunk เพื่อประหยัดแรม
        
        เนื่องจาก pandas.read_excel() ไม่รองรับ chunksize
        เราจึงใช้ openpyxl อ่านข้อมูลทีละ chunk แล้วแปลงเป็น DataFrame
        
        Args:
            path: ที่อยู่ไฟล์ Excel
            sheet_name: ชื่อ Sheet ที่ต้องการอ่าน
            chunk_size: จำนวนแถวต่อ chunk (แนะนำ 10,000)
            usecols: รายชื่อคอลัมน์ที่ต้องการอ่าน (optional, สำหรับเพิ่มความเร็ว)
            
        Yields:
            DataFrame แต่ละ chunk
            
        Example:
            >>> reader = PandasSheetReader()
            >>> for chunk in reader.read_sheet_chunked(path, sheet, 10000):
            ...     process(chunk)  # ประมวลผลทีละ 10,000 แถว
        """
        import pandas as pd
        from openpyxl import load_workbook
        
        try:
            # โหลดไฟล์แบบ read_only เพื่อประหยัดแรม
            wb = load_workbook(path.value, read_only=True, data_only=True)
            ws = wb[sheet_name.value]
            
            # หา header row ที่ถูกต้อง
            header_row_idx = self._find_header_row_openpyxl(ws)
            
            # อ่าน header
            rows_iter = ws.iter_rows(values_only=True)
            header = None
            
            # Skip ไปยัง header row
            for i, row in enumerate(rows_iter):
                if i == header_row_idx:
                    header = row
                    break
            
            if header is None:
                raise ValueError("Could not find header row")
            
            # ทำความสะอาด header
            header = self._clean_header_tuple(header)
            
            # ถ้ามี usecols ให้หา index ของคอลัมน์ที่ต้องการ
            if usecols:
                col_indices = [i for i, col in enumerate(header) if col in usecols]
                filtered_header = [header[i] for i in col_indices]
            else:
                col_indices = None
                filtered_header = header
            
            # อ่านข้อมูลทีละ chunk
            chunk_data = []
            for row in rows_iter:
                # ถ้ามี usecols ให้เลือกเฉพาะคอลัมน์ที่ต้องการ
                if col_indices:
                    filtered_row = [row[i] if i < len(row) else None for i in col_indices]
                else:
                    filtered_row = row
                
                chunk_data.append(filtered_row)
                
                # เมื่อครบ chunk_size ให้ yield DataFrame
                if len(chunk_data) >= chunk_size:
                    df = pd.DataFrame(chunk_data, columns=filtered_header)
                    # แปลงทุก column เป็น string
                    df = df.astype(str)
                    yield df
                    chunk_data = []  # เคลียร์ chunk
            
            # yield chunk สุดท้าย (ถ้ามีเหลือ)
            if chunk_data:
                df = pd.DataFrame(chunk_data, columns=filtered_header)
                df = df.astype(str)
                yield df
            
            wb.close()
                
        except Exception as e:
            raise ValueError(
                f"Error reading sheet '{sheet_name.value}' in chunks: {e}"
            )
    
    def _find_header_row_openpyxl(self, worksheet) -> int:
        """
        หา header row ที่ถูกต้องจาก openpyxl worksheet
        """
        max_valid_cols = 0
        best_row = 0
        
        # ตรวจสอบ 7 แถวแรก
        for row_idx, row in enumerate(worksheet.iter_rows(max_row=7, values_only=True)):
            # นับจำนวน column ที่ valid
            valid_cols = [cell for cell in row if cell is not None and str(cell).strip() != '']
            unique_count = len(set(valid_cols))
            
            if unique_count > max_valid_cols:
                max_valid_cols = unique_count
                best_row = row_idx
        
        return best_row
    
    def _clean_header_tuple(self, header):
        """
        ทำความสะอาด header tuple จาก openpyxl
        """
        new_header = []
        seen = {}
        
        for idx, col in enumerate(header):
            # แปลงเป็น string
            col_str = str(col) if col is not None else f'Column_{idx}'
            
            # ถ้าเป็น None หรือว่าง ให้ตั้งชื่อใหม่
            if col is None or col_str.strip() == '' or col_str == 'None':
                col_str = f'Column_{idx}'
            
            # จัดการ column ที่ซ้ำกัน
            if col_str in seen:
                seen[col_str] += 1
                col_str = f'{col_str}_{seen[col_str]}'
            else:
                seen[col_str] = 0
            
            new_header.append(col_str)
        
        return new_header
    
    def estimate_row_count(self, path: FilePath, sheet_name: SheetName) -> int:
        """
        ประมาณการจำนวนแถวโดยไม่โหลดข้อมูลทั้งหมด
        
        ใช้ openpyxl ในโหมด read_only เพื่อความเร็ว
        วิธีนี้เร็วกว่าการโหลดข้อมูลทั้งหมดมาก
        
        Args:
            path: ที่อยู่ไฟล์ Excel
            sheet_name: ชื่อ Sheet
            
        Returns:
            จำนวนแถวโดยประมาณ (รวม header)
            
        Example:
            >>> reader = PandasSheetReader()
            >>> count = reader.estimate_row_count(path, sheet)
            >>> print(f"File has approximately {count} rows")
        """
        try:
            from openpyxl import load_workbook
            
            # โหลดแบบ read_only เพื่อความเร็ว
            wb = load_workbook(path.value, read_only=True, data_only=True)
            ws = wb[sheet_name.value]
            
            # max_row ให้จำนวนแถวโดยไม่ต้องอ่านข้อมูล
            row_count = ws.max_row
            
            wb.close()
            return row_count
            
        except Exception as e:
            # ถ้า estimate ไม่ได้ ให้ return 0 แทนที่จะ error
            # เพราะนี่เป็นแค่การประมาณการ
            return 0